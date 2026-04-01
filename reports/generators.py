"""Generadores de reportes en Markdown, Excel y PDF."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from shared.errors import ReportGenerationError
from reports.types import ReportRequest


def generate_markdown(request: ReportRequest) -> str:
    """Genera un reporte en formato Markdown."""
    # Si hay raw markdown y no hay sections, usar directamente
    if request.markdown and not request.sections:
        return request.markdown

    parts = [f"# {request.title}", ""]

    for section in request.sections:
        parts.append(f"## {section.heading}")
        parts.append("")
        if section.content:
            parts.append(section.content)
            parts.append("")
        if section.table:
            # Tabla pipe-delimited
            headers = section.table.headers
            parts.append("| " + " | ".join(headers) + " |")
            parts.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in section.table.rows:
                cells = [str(c) for c in row]
                parts.append("| " + " | ".join(cells) + " |")
            parts.append("")

    return "\n".join(parts)


def generate_excel(request: ReportRequest, output_path: Path) -> Path:
    """Genera un reporte en formato Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ReportGenerationError(
            "openpyxl no instalado. Ejecuta: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Titulo en A1 bold
    ws["A1"] = request.title
    ws["A1"].font = Font(bold=True, size=14)
    current_row = 3

    for section in request.sections:
        ws.cell(row=current_row, column=1, value=section.heading).font = Font(
            bold=True, size=12
        )
        current_row += 1
        if section.content:
            ws.cell(row=current_row, column=1, value=section.content)
            current_row += 1
        current_row += 1

        # Si la seccion tiene tabla, crear hoja adicional
        if section.table:
            safe_name = section.heading[:31]  # Max 31 chars para nombre de hoja
            ts = wb.create_sheet(title=safe_name)
            # Headers
            for col_idx, header in enumerate(section.table.headers, 1):
                cell = ts.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            # Datos
            for row_idx, row in enumerate(section.table.rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ts.cell(row=row_idx, column=col_idx, value=value)
            # Auto-ajuste ancho
            for col_idx, header in enumerate(section.table.headers, 1):
                max_len = len(str(header))
                for row in section.table.rows:
                    if col_idx - 1 < len(row):
                        max_len = max(max_len, len(str(row[col_idx - 1])))
                ts.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = min(max_len + 2, 50)

    wb.save(str(output_path))
    return output_path


def _sanitize_text_for_pdf(text: str) -> str:
    """Reemplaza emojis y caracteres no-Latin por equivalentes ASCII.

    Solo se usa como fallback cuando no hay fuente Unicode disponible.
    """
    import re
    _EMOJI_REPLACEMENTS = {
        "\u2705": "[OK]",    # ✅
        "\u274c": "[X]",     # ❌
        "\u26a0": "[!]",     # ⚠
        "\u2139": "[i]",     # ℹ
        "\u2022": "-",       # •
        "\u2013": "-",       # –
        "\u2014": "--",      # —
        "\u2018": "'",       # '
        "\u2019": "'",       # '
        "\u201c": '"',       # "
        "\u201d": '"',       # "
        "\u2026": "...",     # …
        "\u2192": "->",      # →
        "\u2190": "<-",      # ←
        "\u2714": "[OK]",    # ✔
        "\u2716": "[X]",     # ✖
        "\u25cf": "*",       # ●
        "\u25cb": "o",       # ○
        "\u2b50": "*",       # ⭐
        "\U0001f534": "[!]", # 🔴
        "\U0001f7e2": "[OK]",# 🟢
        "\U0001f7e1": "[~]", # 🟡
    }
    for char, repl in _EMOJI_REPLACEMENTS.items():
        text = text.replace(char, repl)
    # Strip remaining non-latin1 characters
    text = re.sub(r'[^\x00-\xff]', '?', text)
    return text


def generate_pdf(request: ReportRequest, output_path: Path) -> Path:
    """Genera un reporte en formato PDF."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise ReportGenerationError(
            "fpdf2 no instalado. Ejecuta: pip install fpdf2"
        )

    import os

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Intentar cargar fuente Unicode en orden de preferencia
    _font_name = "Helvetica"
    _unicode_font = False
    _unicode_candidates = [
        ("ArialUnicode", [
            "/Library/Fonts/Arial Unicode.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        ]),
        ("DejaVu", [
            "DejaVuSansCondensed.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf",
        ]),
    ]
    for name, paths in _unicode_candidates:
        for path in paths:
            if os.path.exists(path):
                try:
                    pdf.add_font(name, "", path)
                    pdf.add_font(name, "B", path)
                    _font_name = name
                    _unicode_font = True
                    break
                except Exception:
                    continue
        if _unicode_font:
            break

    def _safe(text: str) -> str:
        """Sanitiza emojis/símbolos problemáticos para PDF.

        Siempre reemplaza emojis comunes — incluso fuentes Unicode
        como Arial Unicode MS no soportan emojis modernos (✅, ❌, etc.).
        """
        return _sanitize_text_for_pdf(text)

    def _set_font(style: str = "", size: int = 11) -> None:
        pdf.set_font(_font_name, style, size)

    # Titulo
    _set_font("B", 18)
    pdf.cell(0, 10, _safe(request.title), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # Si hay raw markdown y no sections
    if request.markdown and not request.sections:
        _set_font("", 11)
        pdf.multi_cell(0, 6, _safe(request.markdown))
        pdf.output(str(output_path))
        return output_path

    for section in request.sections:
        _set_font("B", 14)
        pdf.cell(0, 10, _safe(section.heading), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        if section.content:
            _set_font("", 11)
            pdf.multi_cell(0, 6, _safe(section.content))
            pdf.ln(3)

        if section.table:
            _set_font("", 9)
            col_count = len(section.table.headers)
            page_width = pdf.w - pdf.l_margin - pdf.r_margin
            col_width = page_width / col_count if col_count else page_width

            # Headers
            _set_font("B", 9)
            for header in section.table.headers:
                pdf.cell(col_width, 7, _safe(str(header)), border=1)
            pdf.ln()

            # Rows
            _set_font("", 9)
            for row in section.table.rows:
                for i, cell in enumerate(row):
                    pdf.cell(col_width, 7, _safe(str(cell)), border=1)
                pdf.ln()
            pdf.ln(5)

    pdf.output(str(output_path))
    return output_path
