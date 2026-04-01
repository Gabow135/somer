"""Tests para reports/generators.py."""

from __future__ import annotations

from pathlib import Path

import pytest

from reports.generators import generate_excel, generate_markdown, generate_pdf
from reports.types import ReportFormat, ReportRequest, ReportSection, TableData


class TestGenerateMarkdown:
    def test_sections_with_content(self) -> None:
        req = ReportRequest(
            title="Mi Reporte",
            sections=[
                ReportSection(heading="Intro", content="Texto de intro"),
                ReportSection(heading="Datos", content="Más texto"),
            ],
        )
        md = generate_markdown(req)
        assert "# Mi Reporte" in md
        assert "## Intro" in md
        assert "Texto de intro" in md
        assert "## Datos" in md

    def test_sections_with_table(self) -> None:
        req = ReportRequest(
            title="Tabla",
            sections=[
                ReportSection(
                    heading="Ventas",
                    content="Detalle:",
                    table=TableData(
                        headers=["Producto", "Total"],
                        rows=[["Widget", "100"], ["Gadget", "200"]],
                    ),
                ),
            ],
        )
        md = generate_markdown(req)
        assert "| Producto | Total |" in md
        assert "| Widget | 100 |" in md
        assert "| --- | --- |" in md

    def test_raw_markdown_fallback(self) -> None:
        req = ReportRequest(
            title="Raw",
            markdown="# Custom\nContenido libre",
        )
        md = generate_markdown(req)
        assert md == "# Custom\nContenido libre"

    def test_sections_take_precedence(self) -> None:
        req = ReportRequest(
            title="Con Sections",
            markdown="# Ignorado",
            sections=[ReportSection(heading="Real", content="Texto real")],
        )
        md = generate_markdown(req)
        assert "## Real" in md
        assert "# Ignorado" not in md

    def test_empty_sections(self) -> None:
        req = ReportRequest(title="Vacío")
        md = generate_markdown(req)
        assert "# Vacío" in md


class TestGenerateExcel:
    def test_creates_file(self, tmp_path: Path) -> None:
        output = tmp_path / "test.xlsx"
        req = ReportRequest(
            title="Excel Test",
            format=ReportFormat.XLSX,
            sections=[
                ReportSection(heading="Resumen", content="Todo bien"),
            ],
        )
        result = generate_excel(req, output)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_with_tables(self, tmp_path: Path) -> None:
        output = tmp_path / "tables.xlsx"
        req = ReportRequest(
            title="Tablas",
            format=ReportFormat.XLSX,
            sections=[
                ReportSection(
                    heading="Datos",
                    table=TableData(
                        headers=["A", "B", "C"],
                        rows=[[1, 2, 3], [4, 5, 6]],
                    ),
                ),
            ],
        )
        result = generate_excel(req, output)
        assert result.exists()

        # Verificar contenido con openpyxl
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(result))
            # Hoja principal
            assert "Reporte" in wb.sheetnames
            # Hoja de datos
            assert "Datos" in wb.sheetnames
            ds = wb["Datos"]
            assert ds.cell(1, 1).value == "A"
            assert ds.cell(1, 2).value == "B"
            assert ds.cell(2, 1).value == 1
            assert ds.cell(3, 3).value == 6
        except ImportError:
            pytest.skip("openpyxl no disponible")


class TestGeneratePdf:
    def test_creates_file(self, tmp_path: Path) -> None:
        output = tmp_path / "test.pdf"
        req = ReportRequest(
            title="PDF Test",
            format=ReportFormat.PDF,
            sections=[
                ReportSection(heading="Intro", content="Contenido del PDF"),
            ],
        )
        try:
            result = generate_pdf(req, output)
            assert result.exists()
            assert result.stat().st_size > 0
        except Exception as exc:
            if "fpdf2" in str(exc) or "fpdf" in str(exc).lower():
                pytest.skip("fpdf2 no disponible")
            raise

    def test_with_raw_markdown(self, tmp_path: Path) -> None:
        output = tmp_path / "raw.pdf"
        req = ReportRequest(
            title="Raw PDF",
            format=ReportFormat.PDF,
            markdown="Contenido markdown crudo sin secciones",
        )
        try:
            result = generate_pdf(req, output)
            assert result.exists()
            assert result.stat().st_size > 0
        except Exception as exc:
            if "fpdf2" in str(exc) or "fpdf" in str(exc).lower():
                pytest.skip("fpdf2 no disponible")
            raise
